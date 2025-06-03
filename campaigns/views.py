from django.shortcuts import render, redirect
import openpyxl
from django.views import View
from django.core.mail import send_mail
from django.conf import settings
from .models import CampaignEmail
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
import requests


def redirect_root_to_campaign(request):
    return redirect("send_campaign")


class EmailCampaignView(View, LoginRequiredMixin):
    def get(self, request):
        return render(request, "campaigns/send-campaign.html")

    def post(self, request):
        excel_file = request.FILES.get("excel-file")
        content = request.POST.get("content")
        subject = request.POST.get("subject")

        if not excel_file or not content:
            messages.error(request, "FILE and CONTENT are required.")
            return redirect("send_campaign")

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        sent_count = 0
        failed_count = 0
        total_count = 0
        failed_emails = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            email = row[0]
            if email:
                total_count += 1
                obj, created = CampaignEmail.objects.get_or_create(email=email)
                if not obj.is_sent:
                    try:
                        send_mail(
                            subject=subject,
                            message=content,
                            from_email=settings.EMAIL_HOST_USER,
                            recipient_list=[email],
                            fail_silently=False,
                        )
                        obj.content = content
                        obj.is_sent = True
                        obj.save()
                        sent_count += 1
                    except Exception as e:
                        failed_count += 1
                        failed_emails.append(email)
                        print(f"Failed to send to {email}: {e}")

        return render(
            request,
            "campaigns/send-campaign.html",
            {
                "total": total_count,
                "sent": sent_count,
                "failed": failed_count,
                "failed_emails": failed_emails,
                "show_result": True,
            },
        )


class WhatsAppTemplateSendView(View):
    ACCESS_TOKEN = ""
    PHONE_NUMBER_ID = ""
    LANGUAGE_CODE = "en_us"

    def get(self, request):
        return render(request, "campaigns/send-whatsapp-message.html")

    def post(self, request):
        excel_file = request.FILES.get("excel-file")
        template_name = request.POST.get("template_name", "").strip()

        if not excel_file or not template_name:
            messages.error(
                request, "Please upload an Excel file and enter the template name."
            )
            return render(request, "campaigns/send-whatsapp-message.html")

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {e}")
            return render(request, "campaigns/send-whatsapp-message.html")

        phone_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            phone = row[0]
            if phone:
                phone_list.append(str(phone).strip())

        url = f"https://graph.facebook.com/v22.0/{self.PHONE_NUMBER_ID}/messages"
        headers = {
            "Authorization": f"Bearer {self.ACCESS_TOKEN}",
            "Content-Type": "application/json",
        }

        success_count = 0
        failed_numbers = []

        for phone in phone_list:
            payload = {
                "messaging_product": "whatsapp",
                "to": phone,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": self.LANGUAGE_CODE},
                },
            }
            response = requests.post(url, headers=headers, json=payload)
            if response.status_code == 200:
                success_count += 1
            else:
                failed_numbers.append((phone, response.json()))

        context = {
            "total": len(phone_list),
            "success": success_count,
            "failed_count": len(failed_numbers),
            "failed_numbers": failed_numbers,
        }
        return render(request, "campaigns/send-whatsapp-message.html", context)
